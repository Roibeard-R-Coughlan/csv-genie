"""
CSV Genie Streamlit app.

Run locally with:
    python -m streamlit run streamlit_app.py
"""

from __future__ import annotations

import csv
from io import BytesIO
import os
from pathlib import Path
import re
from datetime import datetime
import zipfile
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from crm_enrichment_tool import (
    AUDIT_COLUMNS,
    BRAVE_SEARCH_API_KEY_ENV,
    DEFAULT_SEARCH_LOCATION,
    audit_columns_only,
    brave_search,
    clean_cell,
    crm_import_columns_only,
    enrich_dataframe,
    extract_irish_phones,
    fetch_page,
    get_domain,
    is_directory_domain,
    root_url,
)


st.set_page_config(page_title="CSV Genie", layout="wide")

st.title("CSV Genie")
st.caption("Verified CRM import enrichment with candidate review for missing websites, phone numbers and emails.")

research_tab, new_leads_tab, sync_tab = st.tabs(["Research & Approval", "Find New Leads", "Sync Excel Workbook"])

# Keep results alive after download clicks/reruns.
if "result_df" not in st.session_state:
    st.session_state.result_df = None
if "source_file_name" not in st.session_state:
    st.session_state.source_file_name = None
if "last_mode" not in st.session_state:
    st.session_state.last_mode = "preview"
if "last_run_settings" not in st.session_state:
    st.session_state.last_run_settings = None
if "source_df" not in st.session_state:
    st.session_state.source_df = None
if "approved_import_df" not in st.session_state:
    st.session_state.approved_import_df = None
if "approval_log_df" not in st.session_state:
    st.session_state.approval_log_df = None
if "approved_download_signature" not in st.session_state:
    st.session_state.approved_download_signature = None
if "new_lead_candidates_df" not in st.session_state:
    st.session_state.new_lead_candidates_df = None
if "approved_new_leads_df" not in st.session_state:
    st.session_state.approved_new_leads_df = None
if "new_lead_download_signature" not in st.session_state:
    st.session_state.new_lead_download_signature = None
if "new_lead_search_meta" not in st.session_state:
    st.session_state.new_lead_search_meta = None
if "seen_new_leads_df" not in st.session_state:
    st.session_state.seen_new_leads_df = pd.DataFrame(columns=NEW_LEAD_COLUMNS) if "NEW_LEAD_COLUMNS" in globals() else pd.DataFrame()
if "new_lead_export_version" not in st.session_state:
    st.session_state.new_lead_export_version = 0
if "synced_workbook_bytes" not in st.session_state:
    st.session_state.synced_workbook_bytes = None
if "synced_workbook_filename" not in st.session_state:
    st.session_state.synced_workbook_filename = None
if "sync_log_df" not in st.session_state:
    st.session_state.sync_log_df = None

with st.sidebar:
    st.header("Settings")
    mode_label = st.radio(
        "Run mode",
        [
            "Preview only - do not change CRM fields",
            "Verified-only export - fill high-confidence blanks",
        ],
        index=0,
    )
    mode = "preview" if mode_label.startswith("Preview") else "verified_only"

    row_limit = st.number_input(
        "Rows to research",
        min_value=1,
        max_value=500,
        value=10,
        step=1,
        help="This limits researched rows only. The exported CSV still keeps all original rows.",
    )

    min_confidence = st.slider(
        "Minimum confidence for verified-only fill",
        min_value=0.50,
        max_value=1.00,
        value=0.80,
        step=0.05,
    )

    st.subheader("Fields to research")
    target_website = st.checkbox("Website", value=True)
    target_phone = st.checkbox("Phone", value=True)
    target_email = st.checkbox("Email", value=False, help="Email is optional because many Irish SMEs use forms or hide email addresses.")
    target_fields = []
    if target_website:
        target_fields.append("Website")
    if target_phone:
        target_fields.append("Phone")
    if target_email:
        target_fields.append("Email")

    st.subheader("Candidate review")
    show_review_rows_only = st.checkbox(
        "Show only researched rows / candidates",
        value=True,
        help="Keeps the audit table focused on rows with proposals, candidate matches or research notes.",
    )
    show_candidate_columns = st.checkbox(
        "Show candidate match columns",
        value=True,
        help="Shows possible matches even when they are not verified enough to auto-fill CRM fields.",
    )

    st.subheader("Search provider")
    search_provider_label = st.selectbox(
        "Web search backend",
        [
            "Brave Search API - recommended",
            "DuckDuckGo - free",
        ],
        index=0,
        help="Brave is recommended when BRAVE_SEARCH_API_KEY is available. DuckDuckGo is the free fallback.",
    )
    if search_provider_label.startswith("Brave"):
        search_provider = "brave"
    else:
        search_provider = "duckduckgo"
    brave_env_key_exists = bool(os.getenv(BRAVE_SEARCH_API_KEY_ENV))
    if brave_env_key_exists:
        st.caption("BRAVE_SEARCH_API_KEY found in environment/.env")
    elif search_provider.startswith("brave"):
        st.warning("Brave Search API selected without BRAVE_SEARCH_API_KEY will fall back to DuckDuckGo.")
    brave_result_count = st.selectbox(
        "Brave web result count",
        [10, 20],
        index=0,
        help="Use 10 by default. Try 20 only when you want a wider candidate set.",
        disabled=not search_provider.startswith("brave"),
    )
    search_location = st.text_input(
        "Search location bias",
        value=DEFAULT_SEARCH_LOCATION,
        help="Keep this local, e.g. Galway, County Galway, Ireland, to reduce irrelevant directory results.",
    )

    st.subheader("Delays")
    default_search_delay = 0.25 if search_provider.startswith("brave") else 1.0
    search_api_delay = st.number_input(
        "Search API delay (seconds)",
        min_value=0.0,
        max_value=10.0,
        value=default_search_delay,
        step=0.25,
        help="Default is 0.25 for Brave and 1.00 for DuckDuckGo.",
    )
    website_fetch_delay = st.number_input(
        "Website/contact-page fetch delay (seconds)",
        min_value=0.0,
        max_value=10.0,
        value=1.0,
        step=0.25,
    )


def build_run_settings(
    *,
    file_name: str | None,
    mode: str,
    row_limit: int,
    min_confidence: float,
    search_api_delay: float,
    website_fetch_delay: float,
    target_fields: list[str],
    search_provider: str,
    search_location: str,
    brave_result_count: int,
) -> dict:
    """Settings that affect the actual enrichment result.

    Display-only controls such as audit-table filters are intentionally excluded,
    so toggling those will not mark a result as stale.
    """
    return {
        "source_file": file_name or "",
        "mode": mode,
        "rows_to_research": int(row_limit),
        "minimum_confidence": round(float(min_confidence), 2),
        "search_api_delay_seconds": round(float(search_api_delay), 2),
        "website_fetch_delay_seconds": round(float(website_fetch_delay), 2),
        "fields_to_research": list(target_fields),
        "search_provider": search_provider,
        "search_location": search_location or DEFAULT_SEARCH_LOCATION,
        "brave_result_count": int(brave_result_count),
    }


def settings_changed(current: dict, previous: dict | None) -> bool:
    if previous is None:
        return False
    return current != previous


def invalidate_generated_downloads() -> None:
    st.session_state.approved_import_df = None
    st.session_state.approval_log_df = None
    st.session_state.approved_download_signature = None


with research_tab:
    uploaded_file = st.file_uploader("Upload a CRM import CSV", type=["csv"])

current_effective_search_provider = search_provider
if search_provider.startswith("brave") and not os.getenv(BRAVE_SEARCH_API_KEY_ENV):
    current_effective_search_provider = "duckduckgo"

current_run_settings = build_run_settings(
    file_name=uploaded_file.name if uploaded_file is not None else None,
    mode=mode,
    row_limit=int(row_limit),
    min_confidence=float(min_confidence),
    search_api_delay=float(search_api_delay),
    website_fetch_delay=float(website_fetch_delay),
    target_fields=target_fields,
    search_provider=current_effective_search_provider,
    search_location=search_location,
    brave_result_count=int(brave_result_count),
)

with research_tab:
    st.info(
        "Recommended test: Preview only, 10 rows, Website first, Brave count 10, 0.25s search delay and 1.00s website fetch delay. "
        "Use Brave Search API when available, with DuckDuckGo as the free fallback. Keep Search location bias set to Galway, County Galway, Ireland. Use the audit CSV for review."
    )
    st.caption(
        "If a 25-row run seems to stall, watch the progress line after clicking Run research. "
        "The audit/CSV download appears only after the run fully completes."
    )


def is_blank_series(series: pd.Series) -> pd.Series:
    return series.isna() | (series.astype(str).str.strip() == "") | (series.astype(str).str.lower().isin(["nan", "none", "null"]))


def build_visible_audit(audit_df: pd.DataFrame, *, review_rows_only: bool, candidate_columns: bool) -> pd.DataFrame:
    visible = audit_df.copy()

    if not candidate_columns:
        visible = visible[[col for col in visible.columns if not col.startswith("Candidate ")]]

    if review_rows_only:
        signal_cols = [
            col
            for col in visible.columns
            if col.startswith("Proposed ")
            or col.startswith("Candidate ")
            or col.startswith("Best Candidate ")
            or col == "Decision Needed"
            or col == "Enrichment Notes"
        ]
        if signal_cols:
            mask = pd.Series(False, index=visible.index)
            for col in signal_cols:
                mask = mask | ~is_blank_series(visible[col])
            visible = visible[mask]
    return visible


def versioned_filename(original_name: str | None, suffix: str, extension: str) -> str:
    path = Path(original_name or "csv_genie_export")
    base = path.stem or "csv_genie_export"
    if suffix == "approved_v0.1":
        match = re.match(r"^(?P<base>.+)_approved_v(?P<major>\d+)\.(?P<minor>\d+)$", base, flags=re.IGNORECASE)
        if match:
            base = match.group("base")
            suffix = f"approved_v{match.group('major')}.{int(match.group('minor')) + 1}"
    return f"{base}_{suffix}{extension}"


def approval_download_signature(source_file_name: str | None, run_settings: dict | None, approval_df: pd.DataFrame) -> str:
    comparable = approval_df.fillna("").astype(str).to_json(orient="split")
    return f"{source_file_name or ''}|{run_settings or {}}|{comparable}"


APPROVAL_REVIEW_COLUMNS = [
    "Company Name",
    "Area",
    "Existing Website",
    "Proposed Website",
    "Best Candidate Website",
    "Best Candidate Confidence",
    "Best Candidate Rejected Reason",
    "Candidate Website 1 Value",
    "Candidate Website 2 Value",
    "Candidate Website 3 Value",
    "Decision Needed",
    "Approve Website?",
    "Approved Website",
    "Approval Note",
]


def nonblank_value(value: object) -> str:
    return clean_cell(value) or ""


def is_blocked_directory_url(url: object) -> bool:
    value = nonblank_value(url)
    return bool(value and is_directory_domain(get_domain(value)))


def build_website_approval_table(audit_df: pd.DataFrame) -> pd.DataFrame:
    if audit_df.empty:
        return pd.DataFrame(columns=["_source_index", *APPROVAL_REVIEW_COLUMNS])

    signal_cols = [
        "Proposed Website",
        "Best Candidate Website",
        "Best Candidate Rejected Reason",
        "Decision Needed",
        "Candidate Website 1 Value",
        "Candidate Website 2 Value",
        "Candidate Website 3 Value",
        "Candidate Website 1 Rejected Reason",
        "Candidate Website 2 Rejected Reason",
        "Candidate Website 3 Rejected Reason",
    ]

    rows = []
    for idx, row in audit_df.iterrows():
        if not any(nonblank_value(row.get(col)) for col in signal_cols):
            continue

        existing_website = nonblank_value(row.get("Website"))
        proposed_website = nonblank_value(row.get("Proposed Website"))
        best_candidate = nonblank_value(row.get("Best Candidate Website"))
        if is_blocked_directory_url(proposed_website):
            proposed_website = ""
        if is_blocked_directory_url(best_candidate):
            best_candidate = ""
        approved_default = proposed_website or best_candidate
        approve_default = bool(proposed_website and not existing_website)
        decision_needed = nonblank_value(row.get("Decision Needed"))
        if decision_needed == "No reliable candidate found":
            decision_needed = "Manual entry needed"

        rows.append(
            {
                "_source_index": idx,
                "Company Name": nonblank_value(row.get("Company Name")),
                "Area": nonblank_value(row.get("Area")),
                "Existing Website": existing_website,
                "Proposed Website": proposed_website,
                "Best Candidate Website": best_candidate,
                "Best Candidate Confidence": nonblank_value(row.get("Best Candidate Confidence")),
                "Best Candidate Rejected Reason": nonblank_value(row.get("Best Candidate Rejected Reason")),
                "Candidate Website 1 Value": nonblank_value(row.get("Candidate Website 1 Value")),
                "Candidate Website 2 Value": nonblank_value(row.get("Candidate Website 2 Value")),
                "Candidate Website 3 Value": nonblank_value(row.get("Candidate Website 3 Value")),
                "Decision Needed": decision_needed,
                "Approve Website?": approve_default,
                "Approved Website": root_url(approved_default) if approved_default else "",
                "Approval Note": "",
            }
        )

    return pd.DataFrame(rows, columns=["_source_index", *APPROVAL_REVIEW_COLUMNS])


def identify_approval_source(row: pd.Series, approved_website: str) -> str:
    approved_root = root_url(approved_website)
    candidates = [
        ("Proposed Website", row.get("Proposed Website")),
        ("Best Candidate Website", row.get("Best Candidate Website")),
        ("Candidate Website 1", row.get("Candidate Website 1 Value")),
        ("Candidate Website 2", row.get("Candidate Website 2 Value")),
        ("Candidate Website 3", row.get("Candidate Website 3 Value")),
    ]
    for label, value in candidates:
        candidate = nonblank_value(value)
        if candidate and root_url(candidate).lower().rstrip("/") == approved_root.lower().rstrip("/"):
            return label
    return "Manual edit"


def append_manual_note(existing_note: object, approved_website: str, approval_note: str) -> str:
    addition = f"CSV Genie approved website: {approved_website}"
    if approval_note:
        addition = f"{addition}. {approval_note}"
    existing = nonblank_value(existing_note)
    return f"{existing}; {addition}" if existing else addition


def generate_approved_outputs(original_df: pd.DataFrame, approval_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    approved_import_df = original_df.copy()
    log_rows = []
    timestamp = datetime.now().isoformat(timespec="seconds")

    if approval_df.empty:
        return approved_import_df, pd.DataFrame(
            columns=["Company Name", "Approved Website", "Approval Note", "Source candidate/proposal used", "Timestamp", "Applied"]
        )

    for _, approval_row in approval_df.iterrows():
        if not bool(approval_row.get("Approve Website?")):
            continue

        source_index = approval_row.get("_source_index")
        if source_index not in approved_import_df.index:
            continue

        approved_website = nonblank_value(approval_row.get("Approved Website"))
        if not approved_website:
            continue
        if is_blocked_directory_url(approved_website):
            log_rows.append(
                {
                    "Company Name": nonblank_value(approval_row.get("Company Name")),
                    "Approved Website": approved_website,
                    "Approval Note": nonblank_value(approval_row.get("Approval Note")),
                    "Source candidate/proposal used": "Rejected manual approval",
                    "Timestamp": timestamp,
                    "Applied": "No - directory/community site rejected",
                }
            )
            continue

        approved_website = root_url(approved_website)
        original_has_website = bool(nonblank_value(approved_import_df.at[source_index, "Website"]))
        applied = False
        if not original_has_website:
            approved_import_df.at[source_index, "Website"] = approved_website
            applied = True
            if "Research Status" in approved_import_df.columns:
                approved_import_df.at[source_index, "Research Status"] = "Manually approved website"
            if "Manual Notes" in approved_import_df.columns:
                approved_import_df.at[source_index, "Manual Notes"] = append_manual_note(
                    approved_import_df.at[source_index, "Manual Notes"],
                    approved_website,
                    nonblank_value(approval_row.get("Approval Note")),
                )

        log_rows.append(
            {
                "Company Name": nonblank_value(approval_row.get("Company Name")),
                "Approved Website": approved_website,
                "Approval Note": nonblank_value(approval_row.get("Approval Note")),
                "Source candidate/proposal used": identify_approval_source(approval_row, approved_website),
                "Timestamp": timestamp,
                "Applied": "Yes" if applied else "No - existing website not overwritten",
            }
        )

    return approved_import_df, pd.DataFrame(log_rows)


NEW_LEAD_CATEGORIES = [
    "Gym/Studio",
    "Physio",
    "Dental",
    "Legal",
    "Accountants/Finance",
    "Hair/Beauty",
    "Accommodation",
    "Estate Agency",
    "Hospitality",
    "Builder",
    "Plumber",
    "Electrician",
    "Trades",
    "Clinic/Health Other",
    "Retail",
    "Other SME",
]

NEW_LEAD_COLUMNS = [
    "Company Name",
    "Category",
    "Area",
    "Website",
    "Phone",
    "Source URL",
    "Suggested Action",
    "Duplicate Status",
    "Matched Existing Company",
    "Matched Existing Website",
    "Matched Existing Phone",
    "Matched Source File",
    "Matched Field",
    "Duplicate Match Confidence",
    "Duplicate Reason",
    "Confidence",
    "Review Notes",
]

if st.session_state.seen_new_leads_df.empty and not list(st.session_state.seen_new_leads_df.columns):
    st.session_state.seen_new_leads_df = pd.DataFrame(columns=NEW_LEAD_COLUMNS)

NEW_LEAD_APPROVAL_COLUMNS = [
    "Approve Lead?",
    *NEW_LEAD_COLUMNS,
    "Approval Note",
]

CATEGORY_COLUMNS = [
    "Business Category",
    "Website Category",
    "Category",
    "Lead Category",
    "Industry",
]

CATEGORY_SEARCH_TEMPLATES = {
    "Dental": [
        "{area} dental clinic official website",
        "{area} dentist practice contact",
        "{area} orthodontist dental care official site",
    ],
    "Physio": [
        "{area} physiotherapy clinic official website",
        "{area} physio practice contact",
        "{area} physical therapy clinic official site",
    ],
    "Accountants/Finance": [
        "{area} accountants official website",
        "{area} tax advisor bookkeeping firm contact",
        "{area} financial advisor accounting practice official site",
    ],
    "Legal": [
        "{area} solicitor official website",
        "{area} law firm contact",
        "{area} legal services official site",
    ],
    "Gym/Studio": [
        "{area} gym studio official website",
        "{area} fitness studio contact",
        "{area} pilates yoga personal training official site",
    ],
    "Hair/Beauty": [
        "{area} hair salon beauty salon official website",
        "{area} barber nails beauty contact",
        "{area} skincare salon official site",
    ],
    "Accommodation": [
        "{area} guesthouse b&b accommodation official website",
        "{area} hotel bed breakfast contact",
        "{area} self catering accommodation official site",
    ],
    "Estate Agency": [
        "{area} estate agent official website",
        "{area} auctioneer property agent contact",
        "{area} letting agent official site",
    ],
    "Hospitality": [
        "{area} restaurant cafe bar official website",
        "{area} takeaway hospitality business contact",
        "{area} pub restaurant official site",
    ],
    "Builder": [
        "{area} builder contractor official website",
        "{area} construction company contact",
        "{area} building contractor official site",
    ],
    "Plumber": [
        "{area} plumber official website",
        "{area} plumbing contractor contact",
        "{area} emergency plumber official site",
    ],
    "Electrician": [
        "{area} electrician official website",
        "{area} electrical contractor contact",
        "{area} registered electrician official site",
    ],
    "Trades": [
        "{area} trades contractor official website",
        "{area} local trades business contact",
        "{area} repair maintenance contractor official site",
    ],
    "Clinic/Health Other": [
        "{area} health clinic official website",
        "{area} private clinic healthcare contact",
        "{area} wellness clinic official site",
    ],
    "Retail": [
        "{area} retail shop official website",
        "{area} local store contact",
        "{area} independent retailer official site",
    ],
    "Other SME": [
        "{area} local business official website",
        "{area} small business contact",
        "{area} SME official site",
    ],
}

NEW_LEAD_REFERENCE_DOMAINS = {
    "gov.ie",
    "citizensinformation.ie",
    "hse.ie",
    "wikipedia.org",
    "irishlocaldent.com",
}

NEW_LEAD_REJECT_DOMAINS = {
    "saolta.ie",
    "orthodontist.ie",
    "irishlocaldent.com",
}

NEW_LEAD_ASSOCIATION_PUBLIC_DOMAINS = {
    "saolta.ie",
    "hse.ie",
    "orthodontist.ie",
}

NON_IRELAND_LOCATION_TERMS = {
    "newfoundland",
    "canada",
    "ontario",
    "australia",
    "united states",
    "usa",
}

BUSINESS_DOMAIN_WORDS = [
    "forster",
    "court",
    "eyre",
    "square",
    "galway",
    "quay",
    "gate",
    "cuddy",
    "king",
    "niall",
    "cronin",
    "rdent",
    "dental",
    "dentists",
    "dentist",
    "clinic",
    "group",
    "physio",
    "physiotherapy",
    "accountants",
    "accountant",
    "solicitors",
    "law",
    "fitness",
    "gym",
    "studio",
]

GENERIC_RESULT_NAMES = {
    "dentist",
    "dentists",
    "clinic",
    "our clinic",
    "group",
    "list of dentists in county",
    "dentist galway",
    "dentist in galway",
    "trusted dentist in galway",
    "your trusted dentist in galway",
    "specialist dental practice in galway city",
    "patient testimonials",
    "your smile deserves a dentist who truly cares",
    "find local dentists",
    "find local dentists in galway",
    "find local dentist",
    "best dentist in galway",
    "best dentists in galway",
}

GENERIC_SEARCH_TITLE_PATTERNS = [
    r"\bfind\s+local\s+dentists?\b",
    r"\bdentists?\s+galway\b",
    r"\bbest\s+dentists?\s+in\s+galway\b",
    r"\btop\s+\d*\s*dentists?\b",
    r"\blist\s+of\s+dentists?\b",
    r"\bdentists?\s+near\s+me\b",
]

ASSOCIATION_PUBLIC_TERMS = {
    "association",
    "society",
    "public body",
    "hospital",
    "hse",
    "saolta",
    "department",
    "university hospital",
    "public dental service",
}

DIRECTORY_REFERENCE_TERMS = {
    "directory",
    "find local",
    "near me",
    "reviews",
    "list of",
    "best of",
    "top rated",
}

DENTAL_CATEGORY_TERMS = {
    "dental",
    "dentist",
    "dentists",
    "orthodontic",
    "orthodontics",
    "orthodontist",
    "periodontic",
    "periodontics",
    "periodontist",
    "practice",
    "clinic",
}

SME_NEGATIVE_TERMS = {
    "journal",
    "article",
    "news",
    "wikipedia",
    "research paper",
    "health service executive",
}


def normalize_lead_text(value: object) -> str:
    text = nonblank_value(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_lead_domain(value: object) -> str:
    text = nonblank_value(value)
    if not text:
        return ""
    if "://" not in text:
        text = f"https://{text}"
    return get_domain(text)


def normalize_lead_phone(value: object) -> str:
    digits = re.sub(r"\D+", "", nonblank_value(value))
    if digits.startswith("00353"):
        digits = "353" + digits[5:]
    if digits.startswith("353"):
        digits = "0" + digits[3:]
    return digits


def display_phone(value: object) -> str:
    phone = normalize_lead_phone(value)
    return phone or nonblank_value(value)


def first_present_value(row: pd.Series, columns: list[str]) -> str:
    for column in columns:
        value = nonblank_value(row.get(column))
        if value:
            return value
    return ""


def normalize_lead_category(value: object) -> str:
    text = normalize_lead_text(value)
    aliases = {
        "accountants finance": "accountants finance",
        "accountant": "accountants finance",
        "accountants": "accountants finance",
        "finance": "accountants finance",
        "solicitors": "legal",
        "solicitor": "legal",
        "law firm": "legal",
        "gym studio": "gym studio",
        "fitness": "gym studio",
        "hair beauty": "hair beauty",
        "bnb": "accommodation",
        "b b": "accommodation",
        "guesthouse": "accommodation",
        "real estate": "estate agency",
        "estate agents": "estate agency",
        "restaurants": "hospitality",
        "restaurant": "hospitality",
        "pub": "hospitality",
        "dental clinics": "dental",
        "dentists": "dental",
        "physiotherapy": "physio",
        "clinic health other": "clinic health other",
    }
    return aliases.get(text, text)


def normalize_existing_sources(existing_sources: list[object]) -> list[tuple[str, pd.DataFrame]]:
    normalized = []
    for source in existing_sources:
        if isinstance(source, tuple) and len(source) == 2:
            normalized.append((str(source[0]), source[1]))
        else:
            normalized.append(("uploaded CSV", source))
    return normalized


def build_existing_lead_index(existing_dfs: list[object]) -> dict[str, set[str]]:
    index = {
        "names": set(),
        "domains": set(),
        "phones": set(),
        "areas": set(),
        "categories": set(),
        "name_area": set(),
        "name_category": set(),
        "domain_category": set(),
        "phone_category": set(),
        "records": [],
    }
    for source_file, df in normalize_existing_sources(existing_dfs):
        for _, row in df.iterrows():
            raw_name = nonblank_value(row.get("Company Name"))
            raw_area = nonblank_value(row.get("Area"))
            raw_website = nonblank_value(row.get("Website"))
            raw_phone = nonblank_value(row.get("Phone"))
            raw_category = first_present_value(row, CATEGORY_COLUMNS)
            name = normalize_lead_text(raw_name)
            area = normalize_lead_text(raw_area)
            domain = normalize_lead_domain(raw_website)
            phone = normalize_lead_phone(raw_phone)
            category = normalize_lead_category(raw_category)
            if name:
                index["names"].add(name)
            if area:
                index["areas"].add(area)
            if category:
                index["categories"].add(category)
            if name or area:
                index["name_area"].add(f"{name}|{area}")
            if name or category:
                index["name_category"].add(f"{name}|{category}")
            if domain or category:
                index["domain_category"].add(f"{domain}|{category}")
            if phone or category:
                index["phone_category"].add(f"{phone}|{category}")
            if domain:
                index["domains"].add(domain)
            if phone:
                index["phones"].add(phone)
            if name or domain or phone:
                index["records"].append(
                    {
                        "name": name,
                        "area": area,
                        "domain": domain,
                        "phone": phone,
                        "category": category,
                        "company": raw_name,
                        "website": raw_website,
                        "phone_display": raw_phone,
                        "source_file": source_file,
                    }
                )
    return index


def infer_company_name_from_result(title: str, category: str, area: str) -> str:
    cleaned = re.sub(r"<[^>]+>", " ", title or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -|")
    for separator in [" | ", " - ", " – ", " — ", ":"]:
        if separator in cleaned:
            cleaned = cleaned.split(separator)[0].strip()
    noise_terms = ["official website", "official site", "contact us"]
    for term in noise_terms:
        if term:
            cleaned = re.sub(re.escape(term), "", cleaned, flags=re.IGNORECASE).strip(" -|,")
    return cleaned or title.strip() or "Unknown business"


def split_domain_label(label: str) -> list[str]:
    remaining = normalize_lead_text(label).replace(" ", "")
    if not remaining:
        return []
    words = []
    while remaining:
        match = ""
        for word in sorted(BUSINESS_DOMAIN_WORDS, key=len, reverse=True):
            if remaining.startswith(word):
                match = word
                break
        if match:
            words.append(match)
            remaining = remaining[len(match) :]
        else:
            token = re.match(r"^[a-z0-9]+?(?=(?:" + "|".join(re.escape(w) for w in BUSINESS_DOMAIN_WORDS) + r")|$)", remaining)
            chunk = token.group(0) if token else remaining
            words.append(chunk)
            remaining = remaining[len(chunk) :]
    return [word for word in words if word]


def company_name_from_domain(domain: str) -> str:
    label = (domain or "").split(".")[0]
    words = split_domain_label(label)
    return " ".join(word.upper() if word == "rdent" else word.capitalize() for word in words) or "Unknown business"


def extract_page_name_signals(url: str) -> list[str]:
    html, _final_url = fetch_page(url, timeout=4)
    if not html:
        return []
    soup = BeautifulSoup(html, "html.parser")
    signals = []
    selectors = [
        ("meta[property='og:site_name']", "content"),
        ("meta[property='og:title']", "content"),
        ("meta[name='twitter:title']", "content"),
    ]
    for selector, attr in selectors:
        node = soup.select_one(selector)
        value = nonblank_value(node.get(attr)) if node else ""
        if value:
            signals.append(value)
    if soup.title and soup.title.string:
        signals.append(soup.title.string)
    h1 = soup.find("h1")
    if h1:
        signals.append(h1.get_text(" ", strip=True))
    return signals


def is_generic_company_name(name: str) -> bool:
    normalized = normalize_lead_text(name)
    if normalized in GENERIC_RESULT_NAMES:
        return True
    words = normalized.split()
    if len(words) <= 1:
        return True
    generic_tokens = {
        "a",
        "who",
        "your",
        "dentist",
        "dentists",
        "dental",
        "clinic",
        "practice",
        "trusted",
        "specialist",
        "galway",
        "city",
        "county",
        "in",
        "patient",
        "testimonials",
        "smile",
        "deserves",
        "truly",
        "cares",
    }
    return bool(words) and set(words).issubset(generic_tokens)


def lead_tokens(value: object) -> list[str]:
    stopwords = {
        "and",
        "the",
        "of",
        "in",
        "ireland",
        "galway",
        "county",
        "city",
        "clinic",
        "clinics",
        "dental",
        "dentist",
        "dentists",
        "practice",
        "practices",
        "periodontics",
        "periodontic",
        "orthodontics",
        "orthodontic",
        "services",
        "service",
        "ltd",
        "limited",
    }
    return [
        token
        for token in normalize_lead_text(value).split()
        if len(token) >= 3 and token not in stopwords
    ]


def distinctive_name_overlap(name: str, existing_name: str) -> tuple[int, float, str]:
    tokens = set(lead_tokens(name))
    existing_tokens = set(lead_tokens(existing_name))
    if not tokens or not existing_tokens:
        return 0, 0.0, ""
    overlap = tokens & existing_tokens
    confidence = len(overlap) / max(min(len(tokens), len(existing_tokens)), 1)
    return len(overlap), confidence, ", ".join(sorted(overlap))


def is_domain_match(domain: str, blocked_domain: str) -> bool:
    return bool(domain and (domain == blocked_domain or domain.endswith("." + blocked_domain)))


def is_rejected_new_lead_domain(domain: str) -> bool:
    return any(is_domain_match(domain, rejected) for rejected in NEW_LEAD_REJECT_DOMAINS)


def is_association_or_public_body(domain: str, result_text: str) -> bool:
    normalized = normalize_lead_text(result_text)
    if any(is_domain_match(domain, rejected) for rejected in NEW_LEAD_ASSOCIATION_PUBLIC_DOMAINS):
        return True
    return any(term in normalized for term in ASSOCIATION_PUBLIC_TERMS)


def is_generic_search_result(title: str, source_url: str, result_text: str) -> bool:
    normalized_title = normalize_lead_text(title)
    if is_generic_company_name(title):
        return True
    if any(re.search(pattern, normalized_title) for pattern in GENERIC_SEARCH_TITLE_PATTERNS):
        return True
    normalized_url = normalize_lead_text(source_url)
    normalized_text = normalize_lead_text(result_text)
    return any(term in normalized_title or term in normalized_url for term in DIRECTORY_REFERENCE_TERMS) and any(
        term in normalized_text for term in {"dentist", "dentists", "dental"}
    )


def has_dental_category_fit(company_name: str, result_text: str) -> bool:
    normalized = normalize_lead_text(" ".join([company_name, result_text]))
    return any(term in normalized.split() for term in DENTAL_CATEGORY_TERMS)


def is_unrelated_or_non_sme_result(category: str, company_name: str, result_text: str) -> bool:
    normalized = normalize_lead_text(result_text)
    if any(term in normalized for term in SME_NEGATIVE_TERMS):
        return True
    if category == "Dental" and not has_dental_category_fit(company_name, result_text):
        return True
    return False


def is_official_business_website(domain: str, website: str, result_text: str) -> bool:
    if not website or not domain:
        return False
    if is_directory_domain(domain) or is_new_lead_reference_domain(domain) or is_rejected_new_lead_domain(domain):
        return False
    if is_association_or_public_body(domain, result_text):
        return False
    return True


def suggested_action_for_candidate(
    *,
    category: str,
    company_name: str,
    domain: str,
    website: str,
    title: str,
    source_url: str,
    result_text: str,
    duplicate_evidence: dict[str, str],
    name_confident: bool,
) -> tuple[str, str, str]:
    duplicate_status = duplicate_evidence["Duplicate Status"]
    if duplicate_status == "Already exists":
        return "Already exists", "Already exists", duplicate_evidence.get("Duplicate Reason", "")
    if duplicate_status == "Possible duplicate":
        return "Possible duplicate", "Possible duplicate", duplicate_evidence.get("Duplicate Reason", "")
    if is_association_or_public_body(domain, result_text):
        return "Reject: association/public body", "Rejected", "Association, society, hospital, HSE or public-service result"
    if is_directory_domain(domain) or is_new_lead_reference_domain(domain) or is_rejected_new_lead_domain(domain):
        return "Reject: directory/community/reference site", "Rejected", "Directory, community, reference or blocked source domain"
    if is_generic_search_result(title, source_url, result_text):
        return "Reject: generic search result", "Rejected", "Generic list/search result rather than a specific business"
    if has_non_ireland_location_signal(result_text) or is_unrelated_or_non_sme_result(category, company_name, result_text):
        return "Reject: unrelated/non-SME result", "Rejected", "Result does not look like a local SME fit for the selected category"
    if not name_confident:
        return "Manual review needed", "Manual review needed", "Business name could not be confidently inferred"
    if not is_official_business_website(domain, website, result_text):
        return "Manual review needed", "Manual review needed", "No clear official business website"
    return "Recommended new lead", "New lead candidate", "Clear new candidate with an official business website"


def is_recommended_high_confidence_lead(row: pd.Series) -> bool:
    return (
        nonblank_value(row.get("Suggested Action")) == "Recommended new lead"
        and nonblank_value(row.get("Duplicate Status")) == "New lead candidate"
        and nonblank_value(row.get("Confidence")) == "High"
        and bool(nonblank_value(row.get("Website")))
    )


def clean_page_signal(value: str, category: str, area: str) -> str:
    cleaned = infer_company_name_from_result(value, category, area)
    for separator in [" - ", " | ", " – ", " — ", ":"]:
        if separator in cleaned:
            parts = [part.strip() for part in cleaned.split(separator) if part.strip()]
            specific = [part for part in parts if not is_generic_company_name(part)]
            if specific:
                return specific[0]
            return parts[0] if parts else cleaned
    return cleaned


def clean_new_lead_company_name(title: str, domain: str, category: str, area: str, url: str) -> tuple[str, bool, str]:
    inferred = infer_company_name_from_result(title, category, area)
    for signal in extract_page_name_signals(url):
        candidate = clean_page_signal(signal, category, area)
        if candidate and not is_generic_company_name(candidate):
            return candidate, True, "Extracted from page title/H1/meta"
    if inferred and not is_generic_company_name(inferred) and len(lead_tokens(inferred)) >= 2:
        return inferred, True, "Extracted from Brave result title"
    domain_name = company_name_from_domain(domain)
    if domain_name != "Unknown business":
        return domain_name, not is_generic_company_name(domain_name), "Extracted from domain"
    if inferred and not is_generic_company_name(inferred):
        return inferred, True, "Extracted from Brave result title"
    return inferred or domain_name, False, "Generic search title; verify business name manually."


def lead_name_similarity(name: str, existing_name: str) -> float:
    tokens = set(name.split())
    existing_tokens = set(existing_name.split())
    if not tokens or not existing_tokens:
        return 0.0
    return len(tokens & existing_tokens) / len(tokens | existing_tokens)


def lead_area_compatible(area: str, existing_area: str) -> bool:
    if not area or not existing_area:
        return True
    if area == existing_area:
        return True
    area_tokens = set(area.split())
    existing_tokens = set(existing_area.split())
    if not area_tokens or not existing_tokens:
        return False
    return area_tokens.issubset(existing_tokens) or existing_tokens.issubset(area_tokens)


def is_new_lead_reference_domain(domain: str) -> bool:
    return any(domain == blocked or domain.endswith("." + blocked) for blocked in NEW_LEAD_REFERENCE_DOMAINS)


def has_non_ireland_location_signal(result_text: str) -> bool:
    normalized = normalize_lead_text(result_text)
    return any(term in normalized for term in NON_IRELAND_LOCATION_TERMS)


def empty_duplicate_evidence(status: str = "New lead candidate", reason: str = "") -> dict[str, str]:
    return {
        "Duplicate Status": status,
        "Matched Existing Company": "",
        "Matched Existing Website": "",
        "Matched Existing Phone": "",
        "Matched Source File": "",
        "Matched Field": "",
        "Duplicate Match Confidence": "",
        "Duplicate Reason": reason,
    }


def duplicate_evidence_from_record(record: dict[str, str], status: str, field: str, confidence: str, reason: str) -> dict[str, str]:
    return {
        "Duplicate Status": status,
        "Matched Existing Company": record.get("company", ""),
        "Matched Existing Website": record.get("website", ""),
        "Matched Existing Phone": record.get("phone_display", ""),
        "Matched Source File": record.get("source_file", ""),
        "Matched Field": field,
        "Duplicate Match Confidence": confidence,
        "Duplicate Reason": reason,
    }


def duplicate_evidence_for_candidate(
    *,
    company_name: str,
    category: str,
    area: str,
    website: str,
    phone: str,
    existing_index: dict[str, set[str]],
) -> dict[str, str]:
    name_key = normalize_lead_text(company_name)
    area_key = normalize_lead_text(area)
    category_key = normalize_lead_category(category)
    domain_key = normalize_lead_domain(website)
    phone_key = normalize_lead_phone(phone)

    for existing in existing_index["records"]:
        if domain_key and domain_key == existing.get("domain"):
            return duplicate_evidence_from_record(existing, "Already exists", "Website domain", "1.00", "Website domain matches existing duplicate index row")
        if phone_key and phone_key == existing.get("phone"):
            return duplicate_evidence_from_record(existing, "Already exists", "Phone", "1.00", "Phone number matches existing duplicate index row")
        if name_key and area_key and name_key == existing.get("name") and area_key == existing.get("area"):
            return duplicate_evidence_from_record(existing, "Already exists", "Company Name + Area", "1.00", "Company name and area match existing duplicate index row")
        if category_key and name_key and name_key == existing.get("name") and category_key == existing.get("category"):
            return duplicate_evidence_from_record(existing, "Possible duplicate", "Company Name + Category", "0.90", "Company name and category match existing duplicate index row")
        same_area = lead_area_compatible(area_key, existing.get("area", ""))
        same_category = not category_key or not existing.get("category") or category_key == existing.get("category")
        similarity = lead_name_similarity(name_key, existing.get("name", ""))
        if same_area and same_category and similarity >= 0.65:
            return duplicate_evidence_from_record(
                existing,
                "Possible duplicate",
                "Company Name",
                f"{similarity:.2f}",
                "Similar company name with compatible area/category",
            )
        overlap_count, overlap_confidence, overlap_terms = distinctive_name_overlap(name_key, existing.get("name", ""))
        if same_area and overlap_count >= 2 and overlap_confidence >= 0.67:
            return duplicate_evidence_from_record(
                existing,
                "Possible duplicate",
                "Distinctive company tokens",
                f"{overlap_confidence:.2f}",
                f"Distinctive company token overlap despite category/website differences: {overlap_terms}",
            )
    if not website and not phone:
        return empty_duplicate_evidence("Manual review needed", "No website or phone found")
    return empty_duplicate_evidence()


def new_lead_queries_for_category(category: str, area: str) -> list[str]:
    templates = CATEGORY_SEARCH_TEMPLATES.get(category, CATEGORY_SEARCH_TEMPLATES["Other SME"])
    seen = set()
    queries = []
    for template in templates:
        query = template.format(area=area).strip()
        if query and query.casefold() not in seen:
            seen.add(query.casefold())
            queries.append(query)
    return queries


def should_hide_existing_duplicate(duplicate_evidence: dict[str, str]) -> bool:
    if duplicate_evidence.get("Duplicate Status") != "Already exists":
        return False
    source_file = nonblank_value(duplicate_evidence.get("Matched Source File")).casefold()
    source_name = Path(source_file).name
    return "session exported new leads" in source_file or source_name.startswith("new_leads_approved")


def build_new_lead_candidates(
    *,
    category: str,
    area: str,
    target_count: int,
    existing_index: dict[str, set[str]],
    brave_api_key: str | None,
) -> pd.DataFrame:
    if not brave_api_key:
        raise ValueError("BRAVE_SEARCH_API_KEY is required for new lead discovery.")

    queries = new_lead_queries_for_category(category, area)
    rows = []
    seen_sources = set()
    seen_name_domains = set()
    api_errors = []
    new_lead_count = 0

    for query in queries:
        if new_lead_count >= target_count:
            break
        results = brave_search(
            query,
            max_results=20,
            api_key=brave_api_key,
            include_locations=False,
        )
        if len(results) == 1 and results[0].title == "SEARCH_ERROR":
            api_errors.append(results[0].snippet or "Brave Search error")
            continue
        for result in results:
            if new_lead_count >= target_count:
                break
            source_url = nonblank_value(result.url)
            if not source_url or source_url in seen_sources:
                continue
            if result.title == "SEARCH_ERROR":
                continue
            domain = get_domain(source_url)
            result_text = " ".join([result.title, result.snippet, source_url])
            blocked_source = is_directory_domain(domain)
            company_name, name_confident, name_reason = clean_new_lead_company_name(result.title, domain, category, area, source_url)
            website = "" if blocked_source else root_url(source_url)
            phone_values = extract_irish_phones(" ".join([result.snippet, result.candidate_phone]))
            phone = phone_values[0] if phone_values else ""
            dedupe_key = f"{normalize_lead_text(company_name)}|{normalize_lead_domain(website) or domain}"
            if dedupe_key in seen_name_domains:
                continue
            seen_sources.add(source_url)
            seen_name_domains.add(dedupe_key)
            duplicate_evidence = duplicate_evidence_for_candidate(
                company_name=company_name,
                category=category,
                area=area,
                website=website,
                phone=phone,
                existing_index=existing_index,
            )
            if should_hide_existing_duplicate(duplicate_evidence):
                continue
            suggested_action, duplicate_status, action_reason = suggested_action_for_candidate(
                category=category,
                company_name=company_name,
                domain=domain,
                website=website,
                title=result.title,
                source_url=source_url,
                result_text=result_text,
                duplicate_evidence=duplicate_evidence,
                name_confident=name_confident,
            )
            duplicate_evidence["Duplicate Status"] = duplicate_status
            if action_reason and not duplicate_evidence.get("Duplicate Reason"):
                duplicate_evidence["Duplicate Reason"] = action_reason
            notes = []
            notes.append(name_reason)
            if blocked_source:
                notes.append("Directory/community source - manual reference only")
            if duplicate_status in {"Already exists", "Possible duplicate"}:
                notes.append("Matched existing CRM duplicate-check index")
            if action_reason:
                notes.append(action_reason)
            if not website:
                notes.append("No official website candidate from source")
            confidence = "Low"
            if suggested_action == "Recommended new lead" and website:
                confidence = "High"
            elif suggested_action == "Manual review needed":
                confidence = "Medium"
            if duplicate_status == "Already exists" or suggested_action.startswith("Reject:") or blocked_source:
                confidence = "Low"
            rows.append(
                {
                    "Company Name": company_name,
                    "Category": category,
                    "Area": area,
                    "Website": website,
                    "Phone": phone,
                    "Source URL": source_url,
                    "Suggested Action": suggested_action,
                    **duplicate_evidence,
                    "Confidence": confidence,
                    "Review Notes": "; ".join(notes),
                }
            )
            if suggested_action == "Recommended new lead":
                new_lead_count += 1

    output = pd.DataFrame(rows, columns=NEW_LEAD_COLUMNS)
    output.attrs["api_errors"] = api_errors
    output.attrs["queries"] = queries
    return output


def new_lead_search_signature(
    *,
    category: str,
    area: str,
    target_count: int,
    existing_upload_names: list[str],
) -> str:
    return "|".join([category, area, str(int(target_count)), ",".join(sorted(existing_upload_names))])


def build_new_lead_approval_table(candidates_df: pd.DataFrame) -> pd.DataFrame:
    if candidates_df.empty:
        return pd.DataFrame(columns=NEW_LEAD_APPROVAL_COLUMNS)
    approval_df = candidates_df.copy()
    approval_df.insert(0, "Approve Lead?", False)
    approval_df["Approval Note"] = ""
    return approval_df[NEW_LEAD_APPROVAL_COLUMNS]


def build_recommended_new_lead_approval_table(candidates_df: pd.DataFrame) -> pd.DataFrame:
    approval_df = build_new_lead_approval_table(candidates_df)
    if approval_df.empty:
        return approval_df
    approval_df["Approve Lead?"] = approval_df.apply(is_recommended_high_confidence_lead, axis=1)
    return approval_df


def generate_approved_new_leads(approval_df: pd.DataFrame) -> pd.DataFrame:
    if approval_df.empty:
        return pd.DataFrame(columns=NEW_LEAD_COLUMNS)
    approved = approval_df[
        approval_df["Approve Lead?"].fillna(False).astype(bool)
    ].copy()
    for column in ["Approve Lead?", "Approval Note"]:
        if column in approved.columns:
            approved = approved.drop(columns=[column])
    return approved[[col for col in NEW_LEAD_COLUMNS if col in approved.columns]]


def read_local_lead_library() -> list[tuple[str, pd.DataFrame]]:
    lead_library_path = Path("lead-library")
    if not lead_library_path.exists() or not lead_library_path.is_dir():
        return []
    sources = []
    for csv_path in sorted(lead_library_path.glob("*.csv")):
        try:
            sources.append((str(csv_path), pd.read_csv(csv_path, dtype=str, keep_default_na=False)))
        except Exception:
            continue
    return sources


def append_seen_new_leads(approved_df: pd.DataFrame) -> None:
    if approved_df.empty:
        return
    current = st.session_state.seen_new_leads_df
    st.session_state.seen_new_leads_df = pd.concat([current, approved_df], ignore_index=True)


def new_leads_approved_filename() -> str:
    version = max(int(st.session_state.new_lead_export_version), 1)
    return f"new_leads_approved_v0.{version}.csv"


SYNC_FIELDS = ["Website", "Phone", "Email", "Research Status", "Manual Notes"]
INVALID_SHEET_NAME_CHARS = set('/\\?*[]:')
WORKBOOK_NAMESPACE = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def normalize_match_value(value: object) -> str:
    return nonblank_value(value).casefold()


def build_match_key(company_name: object, area: object) -> tuple[str, str]:
    return (normalize_match_value(company_name), normalize_match_value(area))


def is_manual_approval_row(row: pd.Series) -> bool:
    marker_text = " ".join(
        [
            nonblank_value(row.get("Research Status")),
            nonblank_value(row.get("Manual Notes")),
        ]
    ).casefold()
    return "manually approved" in marker_text or "csv genie approved website" in marker_text


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def has_invalid_sheet_name_chars(sheet_name: str) -> bool:
    return any(char in INVALID_SHEET_NAME_CHARS for char in sheet_name)


def sanitise_sheet_name(sheet_name: str) -> str:
    cleaned = "".join(" - " if char in INVALID_SHEET_NAME_CHARS else char for char in sheet_name)
    cleaned = " ".join(cleaned.split()).strip()
    return (cleaned or "Sheet")[:31]


def unique_sheet_name(sheet_name: str, used_names: set[str]) -> str:
    base = sheet_name[:31] or "Sheet"
    candidate = base
    suffix_number = 2
    while candidate.casefold() in used_names:
        suffix = f" {suffix_number}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        suffix_number += 1
    used_names.add(candidate.casefold())
    return candidate


def sanitise_workbook_sheet_names(workbook_bytes: bytes) -> tuple[bytes, list[dict[str, str]]]:
    """Return workbook bytes with invalid sheet names repaired in workbook.xml."""
    try:
        source = BytesIO(workbook_bytes)
        with zipfile.ZipFile(source, "r") as zin:
            workbook_xml = zin.read("xl/workbook.xml")
            root = ET.fromstring(workbook_xml)
            sheets = root.findall("main:sheets/main:sheet", WORKBOOK_NAMESPACE)
            used_names = {
                sheet.attrib.get("name", "").casefold()
                for sheet in sheets
                if not has_invalid_sheet_name_chars(sheet.attrib.get("name", ""))
            }
            changes: list[dict[str, str]] = []
            for sheet in sheets:
                original_name = sheet.attrib.get("name", "")
                if has_invalid_sheet_name_chars(original_name):
                    new_name = unique_sheet_name(sanitise_sheet_name(original_name), used_names)
                    sheet.set("name", new_name)
                    invalid_chars = "".join(char for char in original_name if char in INVALID_SHEET_NAME_CHARS)
                    changes.append(
                        {
                            "original": original_name,
                            "sanitised": new_name,
                            "invalid_chars": invalid_chars,
                        }
                    )

            if not changes:
                return workbook_bytes, []

            ET.register_namespace("", WORKBOOK_NAMESPACE["main"])
            updated_workbook_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output = BytesIO()
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = updated_workbook_xml if item.filename == "xl/workbook.xml" else zin.read(item.filename)
                    zout.writestr(item, data)
            return output.getvalue(), changes
    except Exception as exc:
        raise ValueError(f"Workbook pre-load sheet-name scan failed: {type(exc).__name__}: {exc}") from exc


def sync_excel_workbook(workbook_bytes: bytes, approved_df: pd.DataFrame) -> tuple[bytes, pd.DataFrame, dict]:
    sanitised_workbook_bytes, sheet_name_changes = sanitise_workbook_sheet_names(workbook_bytes)
    try:
        workbook = load_workbook(BytesIO(sanitised_workbook_bytes))
    except Exception as exc:
        invalid_note = ""
        if sheet_name_changes:
            changed_names = "; ".join(f"{change['original']} -> {change['sanitised']}" for change in sheet_name_changes)
            invalid_note = f" Invalid sheet name(s) detected and sanitised before load: {changed_names}."
        raise ValueError(f"Workbook load failed: {type(exc).__name__}: {exc}.{invalid_note}") from exc
    if "CRM Import" not in workbook.sheetnames:
        raise ValueError('Workbook is missing required sheet "CRM Import"')

    sheet_names_before = list(workbook.sheetnames)
    ws = workbook["CRM Import"]
    headers = {
        cell_text(cell.value): column_index
        for column_index, cell in enumerate(ws[1], start=1)
        if cell_text(cell.value)
    }

    required_match_cols = ["Company Name", "Area"]
    missing_match_cols = [col for col in required_match_cols if col not in headers or col not in approved_df.columns]
    if missing_match_cols:
        raise ValueError(f"Missing required sync columns: {', '.join(missing_match_cols)}")

    workbook_rows: dict[tuple[str, str], int] = {}
    duplicate_keys = set()
    for row_number in range(2, ws.max_row + 1):
        key = build_match_key(
            ws.cell(row=row_number, column=headers["Company Name"]).value,
            ws.cell(row=row_number, column=headers["Area"]).value,
        )
        if not any(key):
            continue
        if key in workbook_rows:
            duplicate_keys.add(key)
            continue
        workbook_rows[key] = row_number

    log_rows = [
        {
            "Company Name": "",
            "Area": "",
            "Field updated": "",
            "Old value": change["original"],
            "New value": change["sanitised"],
            "Status": f"Sheet name sanitised before workbook load; invalid character(s): {change['invalid_chars']}",
        }
        for change in sheet_name_changes
    ]
    updates_applied = 0
    updated_row_numbers = set()

    for _, approved_row in approved_df.iterrows():
        key = build_match_key(approved_row.get("Company Name"), approved_row.get("Area"))
        company_name = nonblank_value(approved_row.get("Company Name"))
        area = nonblank_value(approved_row.get("Area"))

        if key not in workbook_rows:
            log_rows.append(
                {
                    "Company Name": company_name,
                    "Area": area,
                    "Field updated": "",
                    "Old value": "",
                    "New value": "",
                    "Status": "No matching workbook row",
                }
            )
            continue
        if key in duplicate_keys:
            log_rows.append(
                {
                    "Company Name": company_name,
                    "Area": area,
                    "Field updated": "",
                    "Old value": "",
                    "New value": "",
                    "Status": "Skipped duplicate workbook match",
                }
            )
            continue

        row_number = workbook_rows[key]
        manual_approval = is_manual_approval_row(approved_row)

        for field_name in SYNC_FIELDS:
            if field_name not in approved_df.columns:
                continue
            if field_name not in headers:
                log_rows.append(
                    {
                        "Company Name": company_name,
                        "Area": area,
                        "Field updated": field_name,
                        "Old value": "",
                        "New value": nonblank_value(approved_row.get(field_name)),
                        "Status": "Skipped missing workbook column",
                    }
                )
                continue

            new_value = nonblank_value(approved_row.get(field_name))
            if not new_value:
                continue

            cell = ws.cell(row=row_number, column=headers[field_name])
            old_value = cell_text(cell.value)
            if old_value == new_value:
                continue

            is_contact_field = field_name in {"Website", "Phone", "Email"}
            can_overwrite_contact = field_name == "Website" and manual_approval
            if is_contact_field and old_value and not can_overwrite_contact:
                log_rows.append(
                    {
                        "Company Name": company_name,
                        "Area": area,
                        "Field updated": field_name,
                        "Old value": old_value,
                        "New value": new_value,
                        "Status": "Skipped non-empty contact field without field-level manual approval",
                    }
                )
                continue

            cell.value = str(new_value)
            if field_name == "Phone":
                cell.number_format = "@"
            updates_applied += 1
            updated_row_numbers.add(row_number)
            status = "Updated"
            if is_contact_field and old_value:
                status = "Updated non-empty contact field from manual approval"
            log_rows.append(
                {
                    "Company Name": company_name,
                    "Area": area,
                    "Field updated": field_name,
                    "Old value": old_value,
                    "New value": new_value,
                    "Status": status,
                }
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    sync_log_df = pd.DataFrame(
        log_rows,
        columns=["Company Name", "Area", "Field updated", "Old value", "New value", "Status"],
    )
    metadata = {
        "sheet_names_before": sheet_names_before,
        "sheet_names_after": list(workbook.sheetnames),
        "updates_applied": updates_applied,
        "rows_updated": len(updated_row_numbers),
        "sheet_name_changes": sheet_name_changes,
    }
    return output.getvalue(), sync_log_df, metadata


with research_tab:
    if uploaded_file is not None:
        df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
        st.session_state.source_df = df.copy()

        st.subheader("Uploaded file preview")
        st.dataframe(df.head(10), use_container_width=True)

        missing_cols = [col for col in ["Company Name", "Area", "Phone", "Email", "Website"] if col not in df.columns]
        if missing_cols:
            st.error(f"Missing required columns: {', '.join(missing_cols)}")
            st.stop()

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Missing websites", int(is_blank_series(df["Website"]).sum()))
        with col2:
            st.metric("Missing phones", int(is_blank_series(df["Phone"]).sum()))
        with col3:
            st.metric("Missing emails", int(is_blank_series(df["Email"]).sum()))

        selected_missing = 0
        for field in target_fields:
            if field in df.columns:
                selected_missing += int(is_blank_series(df[field]).sum())
        if selected_missing and int(row_limit) > selected_missing:
            st.info(f"Only {selected_missing} blank selected fields are currently available to research. A lower row limit may be faster.")

        if not target_fields:
            st.warning("Select at least one field to research.")
            st.stop()

        if st.button("Run research", type="primary"):
            brave_key = os.getenv(BRAVE_SEARCH_API_KEY_ENV)
            effective_search_provider = search_provider
            if search_provider.startswith("brave") and not brave_key:
                st.warning("Brave Search API is selected but BRAVE_SEARCH_API_KEY is missing. Falling back to DuckDuckGo for this run.")
                effective_search_provider = "duckduckgo"

            progress_bar = st.progress(0, text="Starting research...")
            status_box = st.empty()

            def update_progress(done: int, total: int, company_name: str) -> None:
                if total <= 0:
                    progress_bar.progress(0, text="No rows need research.")
                    return
                pct = min(max(done / total, 0.0), 1.0)
                progress_bar.progress(pct, text=f"Researching {done}/{total}: {company_name}")
                status_box.caption(f"Current/last row: {company_name}")

            with st.spinner("Researching records. This may take a few minutes. Keep this tab open..."):
                result_df = enrich_dataframe(
                    df,
                    mode=mode,
                    min_confidence=float(min_confidence),
                    row_limit=int(row_limit),
                    search_api_delay=float(search_api_delay),
                    website_fetch_delay=float(website_fetch_delay),
                    use_apollo=False,
                    apollo_api_key=None,
                    target_fields=target_fields,
                    progress_callback=update_progress,
                    search_provider=effective_search_provider,
                    brave_api_key=brave_key,
                    serpapi_api_key=None,
                    search_location=search_location,
                    brave_result_count=int(brave_result_count),
                )
            progress_bar.progress(1.0, text="Research completed.")
            status_box.empty()

            st.session_state.result_df = result_df
            st.session_state.source_file_name = uploaded_file.name
            st.session_state.last_mode = mode
            st.session_state.last_run_settings = build_run_settings(
                file_name=uploaded_file.name,
                mode=mode,
                row_limit=int(row_limit),
                min_confidence=float(min_confidence),
                search_api_delay=float(search_api_delay),
                website_fetch_delay=float(website_fetch_delay),
                target_fields=target_fields,
                search_provider=effective_search_provider,
                search_location=search_location,
                brave_result_count=int(brave_result_count),
            )
            invalidate_generated_downloads()
            st.session_state.synced_workbook_bytes = None
            st.session_state.synced_workbook_filename = None
            st.session_state.sync_log_df = None
            st.success("Research completed.")

    if st.session_state.result_df is not None:
        result_df = st.session_state.result_df
        is_stale_result = settings_changed(current_run_settings, st.session_state.last_run_settings)

        if st.session_state.last_run_settings:
            with st.expander("Last run settings", expanded=False):
                st.json(st.session_state.last_run_settings)

        if is_stale_result:
            if st.session_state.approved_import_df is not None:
                invalidate_generated_downloads()
            st.warning(
                "The table and downloads below are from a previous run. "
                "Your current sidebar settings are different. Run research again before downloading/importing."
            )
            if st.button("Clear old results"):
                st.session_state.result_df = None
                st.session_state.last_run_settings = None
                invalidate_generated_downloads()
                st.session_state.synced_workbook_bytes = None
                st.session_state.synced_workbook_filename = None
                st.session_state.sync_log_df = None
                st.rerun()

        audit_df = audit_columns_only(result_df)
        final_import_df = crm_import_columns_only(result_df)
        visible_audit = build_visible_audit(
            audit_df,
            review_rows_only=show_review_rows_only,
            candidate_columns=show_candidate_columns,
        )

        proposed_website_count = 0 if "Proposed Website" not in audit_df.columns else int((~is_blank_series(audit_df["Proposed Website"])).sum())
        strong_candidate_count = 0
        if "Best Candidate Confidence" in audit_df.columns:
            best_conf = pd.to_numeric(audit_df["Best Candidate Confidence"], errors="coerce").fillna(0)
            strong_candidate_count = int(((best_conf >= 0.65) & is_blank_series(audit_df.get("Proposed Website", pd.Series("", index=audit_df.index)))).sum())
        candidate_website_count = int(
            sum((~is_blank_series(audit_df[col])).sum() for col in audit_df.columns if col.startswith("Candidate Website") and col.endswith("Value"))
        )
        candidate_phone_count = int(
            sum((~is_blank_series(audit_df[col])).sum() for col in audit_df.columns if col.startswith("Candidate Phone") and col.endswith("Value"))
        )

        st.subheader("Audit review")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Verified website proposals", proposed_website_count)
        c2.metric("Strong website candidates", strong_candidate_count)
        c3.metric(
            "Candidate URLs found",
            candidate_website_count,
            help="Candidate URLs are research results only. They are not written to the CRM unless verified or manually approved.",
        )
        c4.metric("Phone candidates", candidate_phone_count)

        if visible_audit.empty:
            st.warning("No proposals, candidates or research notes are visible with the current filters. Untick the sidebar filter to show all rows.")
        else:
            st.dataframe(visible_audit, use_container_width=True)

        st.subheader("Manual Approval")
        approval_table = build_website_approval_table(audit_df)
        if approval_table.empty:
            st.info("No website proposals or candidates are available for approval yet.")
        else:
            edited_approval_df = st.data_editor(
                approval_table,
                use_container_width=True,
                hide_index=True,
                disabled=[
                    col
                    for col in approval_table.columns
                    if col not in {"Approve Website?", "Approved Website", "Approval Note"}
                ],
                column_config={
                    "_source_index": None,
                    "Approve Website?": st.column_config.CheckboxColumn("Approve Website?"),
                    "Approved Website": st.column_config.TextColumn("Approved Website"),
                    "Approval Note": st.column_config.TextColumn("Approval Note"),
                },
                key="website_approval_editor",
            )
            current_approval_signature = approval_download_signature(
                st.session_state.source_file_name,
                st.session_state.last_run_settings,
                edited_approval_df,
            )
            if (
                st.session_state.approved_import_df is not None
                and st.session_state.approved_download_signature != current_approval_signature
            ):
                invalidate_generated_downloads()

            if st.button("Generate approved CRM import CSV", disabled=is_stale_result or st.session_state.source_df is None):
                approved_import_df, approval_log_df = generate_approved_outputs(
                    st.session_state.source_df,
                    edited_approval_df,
                )
                st.session_state.approved_import_df = approved_import_df
                st.session_state.approval_log_df = approval_log_df
                st.session_state.approved_download_signature = current_approval_signature
                applied_count = 0 if approval_log_df.empty else int((approval_log_df["Applied"] == "Yes").sum())
                st.success(f"Approved CRM import CSV generated. Websites applied to {applied_count} blank Website field(s).")

            if (
                st.session_state.approved_import_df is not None
                and st.session_state.approved_download_signature == current_approval_signature
            ):
                st.download_button(
                    "Download approved CRM import CSV for CRM upload",
                    data=st.session_state.approved_import_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
                    file_name=versioned_filename(st.session_state.source_file_name, "approved_v0.1", ".csv"),
                    disabled=is_stale_result,
                    mime="text/csv",
                    key="download_approved_crm_csv",
                )

        with st.expander("Advanced downloads", expanded=False):
            st.download_button(
                "Download audit CSV for review - not for CRM import",
                data=audit_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
                file_name="csv_genie_audit.csv",
                disabled=is_stale_result,
                mime="text/csv",
                key="download_audit_csv",
            )

            if st.session_state.approval_log_df is not None:
                st.download_button(
                    "Download approval log CSV",
                    data=st.session_state.approval_log_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
                    file_name="csv_genie_approval_log.csv",
                    disabled=is_stale_result,
                    mime="text/csv",
                    key="download_approval_log_csv",
                )

            if st.session_state.last_mode == "preview":
                st.warning("Preview mode keeps CRM fields unchanged. The audit CSV is the useful file at this stage; the CRM import CSV is disabled to avoid importing unchanged data by mistake.")
                st.download_button(
                    "Download CRM import CSV - disabled in Preview mode",
                    data=final_import_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
                    file_name="csv_genie_crm_import_preview_unchanged.csv",
                    mime="text/csv",
                    disabled=True,
                    key="download_crm_disabled",
                )
            else:
                st.warning("Verified-only mode fills only blank fields that meet the confidence threshold. Still review the audit CSV before importing.")
                st.download_button(
                    "Download CRM import CSV for website import",
                    data=final_import_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
                    file_name="csv_genie_crm_import.csv",
                    mime="text/csv",
                    disabled=is_stale_result,
                    key="download_crm_csv",
                )


with new_leads_tab:
    st.info("For repeat runs, upload your latest approved CRM file and previous new_leads_approved files as duplicate indexes.")
    existing_crm_uploads = st.file_uploader(
        "Upload existing CRM CSVs for duplicate check",
        type=["csv"],
        accept_multiple_files=True,
        key="new_leads_existing_crm_csvs",
    )
    use_local_lead_library = st.checkbox(
        "Also use local ignored lead-library/ folder as duplicate index",
        value=False,
        help="Optional. Reads local CSVs from lead-library/ for duplicate checks only; files remain read-only and ignored by Git.",
    )
    show_only_new_leads = st.checkbox(
        "Show only new non-duplicate leads",
        value=True,
    )
    new_lead_col1, new_lead_col2, new_lead_col3 = st.columns([2, 2, 1])
    with new_lead_col1:
        new_lead_category = st.selectbox("Category", NEW_LEAD_CATEGORIES, index=NEW_LEAD_CATEGORIES.index("Dental"))
    with new_lead_col2:
        new_lead_area = st.text_input("Area", value="Galway", key="new_lead_area")
    with new_lead_col3:
        new_lead_target_count = st.number_input("Target new non-duplicate leads", min_value=1, max_value=100, value=10, step=1)

    existing_sources = []
    for existing_upload in existing_crm_uploads:
        try:
            existing_sources.append((existing_upload.name, pd.read_csv(existing_upload, dtype=str, keep_default_na=False)))
        except Exception as exc:
            st.warning(f"{existing_upload.name} could not be read: {type(exc).__name__}: {exc}")
    if use_local_lead_library:
        local_sources = read_local_lead_library()
        existing_sources.extend(local_sources)
        st.caption(f"Local lead-library CSVs indexed: {len(local_sources)}")
    if not st.session_state.seen_new_leads_df.empty:
        existing_sources.append(("Session exported new leads", st.session_state.seen_new_leads_df))
    duplicate_index = build_existing_lead_index(existing_sources)
    st.metric("Existing duplicate-index rows", sum(len(df) for _, df in existing_sources))
    new_lead_signature = new_lead_search_signature(
        category=new_lead_category,
        area=new_lead_area,
        target_count=int(new_lead_target_count),
        existing_upload_names=[name for name, _df in existing_sources],
    )
    if (
        st.session_state.new_lead_search_meta
        and st.session_state.new_lead_search_meta.get("signature") != new_lead_signature
    ):
        st.session_state.new_lead_candidates_df = None
        st.session_state.approved_new_leads_df = None
        st.session_state.new_lead_download_signature = None
        st.session_state.new_lead_search_meta = None

    if st.button("Find new lead candidates", type="primary", disabled=not os.getenv(BRAVE_SEARCH_API_KEY_ENV)):
        try:
            with st.spinner("Searching Brave for candidate businesses..."):
                candidates_df = build_new_lead_candidates(
                    category=new_lead_category,
                    area=new_lead_area,
                    target_count=int(new_lead_target_count),
                    existing_index=duplicate_index,
                    brave_api_key=os.getenv(BRAVE_SEARCH_API_KEY_ENV),
                )
                st.session_state.new_lead_candidates_df = candidates_df
                st.session_state.approved_new_leads_df = None
                st.session_state.new_lead_download_signature = None
                st.session_state.new_lead_search_meta = {
                    "signature": new_lead_signature,
                    "api_errors": list(candidates_df.attrs.get("api_errors", [])),
                    "queries": list(candidates_df.attrs.get("queries", [])),
                }
            st.success(f"New Lead Candidates CSV ready with {len(st.session_state.new_lead_candidates_df)} candidate row(s).")
        except Exception as exc:
            st.error(f"New lead search failed: {type(exc).__name__}: {exc}")

    if not os.getenv(BRAVE_SEARCH_API_KEY_ENV):
        st.warning("BRAVE_SEARCH_API_KEY is required for Find New Leads.")

    if st.session_state.new_lead_candidates_df is not None:
        search_meta = st.session_state.new_lead_search_meta or {}
        api_errors = search_meta.get("api_errors", [])
        status_counts = st.session_state.new_lead_candidates_df["Duplicate Status"].value_counts().to_dict()
        action_counts = st.session_state.new_lead_candidates_df["Suggested Action"].value_counts().to_dict()
        visible_new_leads_df = st.session_state.new_lead_candidates_df
        if show_only_new_leads:
            visible_new_leads_df = visible_new_leads_df[visible_new_leads_df["Duplicate Status"] == "New lead candidate"]
        metric_cols = st.columns(5)
        metric_cols[0].metric("Recommended", int(action_counts.get("Recommended new lead", 0)))
        metric_cols[1].metric("Possible duplicates", int(status_counts.get("Possible duplicate", 0)))
        metric_cols[2].metric("Already exists", int(status_counts.get("Already exists", 0)))
        metric_cols[3].metric("Manual/rejected", int(action_counts.get("Manual review needed", 0)) + sum(count for action, count in action_counts.items() if action.startswith("Reject:")))
        metric_cols[4].metric("API errors", len(api_errors))
        if api_errors:
            with st.expander("New lead search API errors", expanded=False):
                for error in api_errors:
                    st.write(error)

        st.dataframe(visible_new_leads_df, use_container_width=True)
        st.subheader("Approve new leads for export")
        select_recommended_only = st.checkbox(
            "Select recommended high-confidence leads only",
            value=False,
            help="Selects only clear new leads with high confidence, official business websites and no duplicate/reject/manual-review flags.",
        )
        if select_recommended_only:
            new_lead_approval_table = build_recommended_new_lead_approval_table(visible_new_leads_df)
        else:
            new_lead_approval_table = build_new_lead_approval_table(visible_new_leads_df)
        editor_signature = re.sub(r"[^a-zA-Z0-9_]+", "_", str(search_meta.get("signature", "current")))[:120]
        edited_new_lead_approval_df = st.data_editor(
            new_lead_approval_table,
            use_container_width=True,
            hide_index=True,
            disabled=[
                col
                for col in new_lead_approval_table.columns
                if col not in {"Approve Lead?", "Company Name", "Area", "Category", "Website", "Phone", "Review Notes", "Approval Note"}
            ],
            column_config={
                "Approve Lead?": st.column_config.CheckboxColumn("Approve Lead?"),
                "Company Name": st.column_config.TextColumn("Company Name"),
                "Area": st.column_config.TextColumn("Area"),
                "Category": st.column_config.TextColumn("Category"),
                "Website": st.column_config.TextColumn("Website"),
                "Phone": st.column_config.TextColumn("Phone"),
                "Suggested Action": st.column_config.TextColumn("Suggested Action"),
                "Review Notes": st.column_config.TextColumn("Review Notes"),
                "Approval Note": st.column_config.TextColumn("Approval Note"),
            },
            key=f"new_lead_approval_editor_{editor_signature}_{int(select_recommended_only)}",
        )
        current_new_lead_download_signature = approval_download_signature(
            "new_lead_candidates.csv",
            search_meta,
            edited_new_lead_approval_df,
        )
        if (
            st.session_state.approved_new_leads_df is not None
            and st.session_state.new_lead_download_signature != current_new_lead_download_signature
        ):
            st.session_state.approved_new_leads_df = None
            st.session_state.new_lead_download_signature = None

        if st.button("Generate approved new leads CSV"):
            st.session_state.approved_new_leads_df = generate_approved_new_leads(edited_new_lead_approval_df)
            st.session_state.new_lead_download_signature = current_new_lead_download_signature
            append_seen_new_leads(st.session_state.approved_new_leads_df)
            st.session_state.new_lead_export_version += 1
            st.success(f"Approved new leads CSV generated with {len(st.session_state.approved_new_leads_df)} row(s).")

        if (
            st.session_state.approved_new_leads_df is not None
            and st.session_state.new_lead_download_signature == current_new_lead_download_signature
        ):
            st.download_button(
                "Download approved new leads CSV",
                data=st.session_state.approved_new_leads_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
                file_name=new_leads_approved_filename(),
                mime="text/csv",
                key="download_approved_new_leads_csv",
            )


with sync_tab:
    excel_upload = st.file_uploader(
        "Upload original Excel workbook",
        type=["xlsx", "xlsm"],
        key="excel_sync_workbook",
    )
    approved_csv_upload = st.file_uploader(
        "Upload approved CRM import CSV",
        type=["csv"],
        key="excel_sync_approved_csv",
    )

    if st.button(
        "Generate synced Excel workbook",
        disabled=excel_upload is None or approved_csv_upload is None,
    ):
        try:
            approved_sync_df = pd.read_csv(approved_csv_upload, dtype=str, keep_default_na=False)
            workbook_bytes, sync_log_df, sync_metadata = sync_excel_workbook(
                excel_upload.getvalue(),
                approved_sync_df,
            )
            st.session_state.synced_workbook_bytes = workbook_bytes
            st.session_state.synced_workbook_filename = versioned_filename(
                excel_upload.name,
                "synced_v0.1",
                ".xlsx",
            )
            st.session_state.sync_log_df = sync_log_df
            st.success(
                "Synced Excel workbook generated. "
                f"Rows updated: {sync_metadata['rows_updated']}. "
                f"Fields updated: {sync_metadata['updates_applied']}. "
                f"Sheets preserved: {len(sync_metadata['sheet_names_after'])}."
            )
            if sync_metadata["sheet_name_changes"]:
                changed_names = "; ".join(
                    f"{change['original']} -> {change['sanitised']}"
                    for change in sync_metadata["sheet_name_changes"]
                )
                st.warning(f"Invalid sheet name(s) detected and sanitised: {changed_names}")
        except Exception as exc:
            st.error(f"Excel sync failed: {type(exc).__name__}: {exc}")

    if st.session_state.synced_workbook_bytes is not None:
        st.download_button(
            "Download synced Excel workbook",
            data=st.session_state.synced_workbook_bytes,
            file_name=st.session_state.synced_workbook_filename or "csv_genie_synced_v0.1.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_synced_workbook",
        )

    if st.session_state.sync_log_df is not None:
        st.download_button(
            "Download sync log CSV",
            data=st.session_state.sync_log_df.to_csv(index=False, quoting=csv.QUOTE_ALL).encode("utf-8-sig"),
            file_name="csv_genie_excel_sync_log.csv",
            mime="text/csv",
            key="download_excel_sync_log",
        )
